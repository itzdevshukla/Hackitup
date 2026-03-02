from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from datetime import datetime
from .models import Event
from dashboard.models import EventAccess
from challenges.models import UserChallenge
from django.db.models import Sum, Min, Max, Count
import openpyxl
import secrets
import string
from django.http import HttpResponse
from dashboard.models import EventAccess, EventRegistration






# ------------------------ADMIN DASHBOARD VIEW---------------------------------------
@login_required(login_url='/accounts/login/')
def admin_dashboard_view(request):
    """Main admin dashboard"""
    
    if not request.user.is_staff:
        messages.error(request, '❌ Access denied!')
        return redirect('dashboard')
    
    # Calculate stats
    total_users = User.objects.count()
    total_events = Event.objects.count()
    live_events = Event.objects.filter(status='live').count()
    upcoming_events = Event.objects.filter(status='upcoming').count()
    
    context = {
        'total_users': total_users,
        'total_events': total_events,
        'live_events': live_events,
        'upcoming_events': upcoming_events,
    }
    
    return render(request, 'administration/admin.html', context)


# ----------------------------ALL USER DETAILS------------------------------------------------------------

@login_required(login_url='/accounts/login/')
def admin_users_view(request):
    """View all participants"""
    
    if not request.user.is_staff:
        messages.error(request, '❌ Access denied!')
        return redirect('dashboard')
    
    # Get all users
    users = User.objects.all().order_by('-date_joined')
    
    
    total_users = users.count()
    active_users = users.filter(is_active=True).count()
    admin_count = users.filter(is_staff=True).count()
    
    
    first_day = datetime.now().replace(day=1, hour=0, minute=0, second=0)
    new_users = users.filter(date_joined__gte=first_day).count()
    
    context = {
        'users': users,
        'total_users': total_users,
        'active_users': active_users,
        'admin_count': admin_count,
        'new_users': new_users,
    }
    
    return render(request, 'administration/users.html', context)



# -----------------------------------------ADMIN USER DELETE----------------------------------------------


@login_required
def admin_delete_user_view(request, user_id):
    """Delete user"""
    
    if not request.user.is_staff:
        messages.error(request, '❌ Access denied!')
        return redirect('dashboard')
    
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        
        
        if user.id == request.user.id:
            messages.error(request, '❌ You cannot delete your own account!')
            return redirect('admin_users')
        
        username = user.username
        user.delete()
        
        messages.success(request, f'✅ User "{username}" deleted successfully!')
        return redirect('admin_users')
    
    return redirect('admin_users')




# ==================================

@login_required(login_url='/accounts/login/')
def admin_user_detail_view(request, user_id):
    """View single user details (placeholder for now)"""
    
    if not request.user.is_staff:
        messages.error(request, '❌ Access denied!')
        return redirect('dashboard')
    
    user = get_object_or_404(User, id=user_id)
    
    # Placeholder - will build detailed view later
    messages.info(request, f'Viewing details for {user.username}')
    return redirect('admin_users')










# Event Managaement -----------------------------------------------------


@login_required(login_url='/accounts/login/')
def admin_manage_events_view(request):
    """Manage/Delete events"""
    
    if not request.user.is_staff:
        messages.error(request, '❌ Access denied!')
        return redirect('dashboard')
    
    # Get all events
    events = Event.objects.all().order_by('-created_at')
    
    # Calculate stats
    total_events = events.count()
    live_events = events.filter(status='live').count()
    upcoming_events = events.filter(status='upcoming').count()
    completed_events = events.filter(status='completed').count()
    
    context = {
        'events': events,
        'total_events': total_events,
        'live_events': live_events,
        'upcoming_events': upcoming_events,
        'completed_events': completed_events,
    }
    
    return render(request, 'administration/admin_manage_events.html', context)





# ADMIN DELETE EVENT=========================
@login_required(login_url='/accounts/login/')
def admin_delete_event_view(request, event_id):
    """Delete event"""
    
    if not request.user.is_staff:
        messages.error(request, '❌ Access denied!')
        return redirect('dashboard')
    
    if request.method == 'POST':
        event = get_object_or_404(Event, id=event_id)
        event_name = event.event_name
        event.delete()
        
        messages.success(request, f'✅ Event "{event_name}" deleted successfully!')
        return redirect('admin_manage_events')
    
    return redirect('admin_manage_events')



# ADMIN ADD EVENT=======================================

from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from administration.models import Event


@login_required(login_url='/accounts/login/')
def admin_add_event_view(request):

    if not request.user.is_staff:
        messages.error(request, '❌ Access denied!')
        return redirect('dashboard')

    new_event_code = None

    if request.method == 'POST':
        try:
            
            start_date = datetime.strptime(
                request.POST.get('start_date'), "%Y-%m-%d"
            ).date()

            start_time = datetime.strptime(
                request.POST.get('start_time'), "%H:%M"
            ).time()

            end_date = datetime.strptime(
                request.POST.get('end_date'), "%Y-%m-%d"
            ).date()

            end_time = datetime.strptime(
                request.POST.get('end_time'), "%H:%M"
            ).time()

            
            reg_start_date = request.POST.get("registration_start_date")
            reg_start_time = request.POST.get("registration_start_time")
            reg_end_date = request.POST.get("registration_end_date")
            reg_end_time = request.POST.get("registration_end_time")

            event = Event.objects.create(
                event_name=request.POST.get('event_name'),
                venue=request.POST.get('venue'),
                description=request.POST.get('description'),
                ctf_type=request.POST.get('ctf_type'),
                max_participants=int(request.POST.get('max_participants')),

                start_date=start_date,
                start_time=start_time,
                end_date=end_date,
                end_time=end_time,

                status=request.POST.get('status'),
                created_by=request.user,
            )

            
            if reg_start_date and reg_start_time:
                event.registration_start_date = datetime.strptime(
                    reg_start_date, "%Y-%m-%d"
                ).date()
                event.registration_start_time = datetime.strptime(
                    reg_start_time, "%H:%M"
                ).time()

            if reg_end_date and reg_end_time:
                event.registration_end_date = datetime.strptime(
                    reg_end_date, "%Y-%m-%d"
                ).date()
                event.registration_end_time = datetime.strptime(
                    reg_end_time, "%H:%M"
                ).time()

            event.save()

            new_event_code = event.access_code
            messages.success(
                request,
                f'✅ Event "{event.event_name}" created successfully!'
            )

        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')

    return render(request, 'administration/admin_add_event.html', {
        'new_event_code': new_event_code
    })



# ----------------------------------------------------------------------------



@login_required(login_url='/accounts/login/')
def admin_settings_view(request):
    """Admin settings (placeholder)"""
    
    if not request.user.is_staff:
        messages.error(request, '❌ Access denied!')
        return redirect('dashboard')
    
    return render(request, 'administration/admin_settings.html')






from django.http import HttpResponseForbidden

from administration.models import Event
from challenges.models import Challenge




# ------------------------------------MANAGE EVENTS-------------------------------------------------
@login_required
def admin_manage_event(request, event_id):

    if not request.user.is_staff:
        return HttpResponseForbidden("Access denied")

    event = get_object_or_404(Event, id=event_id)

    
    challenges = Challenge.objects.filter(event=event)

    
    total_challenges = challenges.count()
    total_points = sum(c.points for c in challenges)

    return render(request, 'administration/manage_individual_event.html', {
        'event': event,
        'challenges': challenges,
        'total_challenges': total_challenges,
        'total_points': total_points,
    })











# -----------------------------EDIT EVENT--------------------------------------------------
from datetime import datetime
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from administration.models import Event


@login_required
def admin_edit_event(request, event_id):

    if not request.user.is_staff:
        messages.error(request, "❌ Access denied")
        return redirect("dashboard")

    event = get_object_or_404(Event, id=event_id)

    if request.method == "POST":
        try:
            
            event.event_name = request.POST.get("event_name")
            event.venue = request.POST.get("venue")
            event.description = request.POST.get("description")
            event.ctf_type = request.POST.get("ctf_type")
            event.max_participants = int(request.POST.get("max_participants"))

            event.start_date = datetime.strptime(
                request.POST.get("start_date"), "%Y-%m-%d"
            ).date()

            event.start_time = datetime.strptime(
                request.POST.get("start_time"), "%H:%M"
            ).time()

            event.end_date = datetime.strptime(
                request.POST.get("end_date"), "%Y-%m-%d"
            ).date()

            event.end_time = datetime.strptime(
                request.POST.get("end_time"), "%H:%M"
            ).time()

            def parse_date(val):
                return datetime.strptime(val, "%Y-%m-%d").date() if val else None

            def parse_time(val):
                return datetime.strptime(val, "%H:%M").time() if val else None

            event.registration_start_date = parse_date(
                request.POST.get("registration_start_date")
            )
            event.registration_start_time = parse_time(
                request.POST.get("registration_start_time")
            )

            event.registration_end_date = parse_date(
                request.POST.get("registration_end_date")
            )
            event.registration_end_time = parse_time(
                request.POST.get("registration_end_time")
            )

            # ===============================
            # STATUS
            # ===============================
            event.status = request.POST.get("status")

            event.save()

            messages.success(
                request,
                f"✅ Event '{event.event_name}' updated successfully"
            )
            return redirect("admin_manage_event", event.id)

        except Exception as e:
            messages.error(request, f"❌ Error: {str(e)}")

    return render(request, "administration/admin_edit_event.html", {
        "event": event
    })




# ----------------------------INDIVIDUAL EVENT USERS-------------------------------------
from django.http import HttpResponseForbidden


@login_required
def admin_event_users(request, event_id):

    if not request.user.is_staff:
        return HttpResponseForbidden("Access denied")

    event = get_object_or_404(Event, id=event_id)

    registrations = EventAccess.objects.filter(event=event).select_related('user')

    return render(request, 'administration/event_users.html', {
        'event': event,
        'registrations': registrations,
        'total_users': registrations.count()
    })












# ================================ADMIN EVENT LEADERBOARD====================================



@login_required
def admin_event_leaderboard(request, event_id):
    if not request.user.is_staff:
        messages.error(request, "Access denied")
        return redirect("dashboard")

    event = get_object_or_404(Event, id=event_id)

    leaderboard = (
        UserChallenge.objects
        .filter(
            challenge__event=event,
            is_correct=True
        )
        .values("user__id", "user__username")
        .annotate(
            total_points=Sum("challenge__points"),
            solves=Count("id"),
            first_solve=Min("submitted_at"),
            last_solve=Max("submitted_at")
        )
        .order_by("-total_points", "last_solve")
    )

    return render(request, "administration/admin_leaderboard.html", {
        "event": event,
        "leaderboard": leaderboard
    })







# ========================USER SUBMISSIONS====================================


@login_required
def admin_user_submissions(request, event_id, user_id):
    if not request.user.is_staff:
        messages.error(request, "Access denied")
        return redirect("dashboard")

    event = get_object_or_404(Event, id=event_id)
    user = get_object_or_404(User, id=user_id)

    access = get_object_or_404(
        EventAccess,
        event=event,
        user=user
    )

    submissions = (
        UserChallenge.objects
        .filter(
            challenge__event=event,
            user=user,
            is_correct=True
        )
        .select_related("challenge")
        .order_by("submitted_at")
    )

    return render(request, "administration/admin_user_submissions.html", {
        "event": event,
        "target_user": user,
        "submissions": submissions,
        "access": access,
    })



# =====================================================================================================
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET
from django.utils import timezone

from challenges.models import UserChallenge
from administration.models import Event


@login_required
@require_GET
def admin_live_correct_submissions_api(request, event_id):

    if not request.user.is_staff:
        return JsonResponse({"error": "Forbidden"}, status=403)

    event = Event.objects.get(id=event_id)

    after_id = request.GET.get("after")
    qs = UserChallenge.objects.filter(
        challenge__event=event,
    ).select_related("user", "challenge").order_by("id")

    if after_id:
        qs = qs.filter(id__gt=after_id)

    data = []
    for s in qs:
        data.append({
            "id": s.id,
            "time": timezone.localtime(s.submitted_at).strftime("%H:%M:%S"),
            "user": s.user.username,
            "challenge": s.challenge.title,
            "flag": s.submitted_flag,
            "is_correct": s.is_correct,
        })

    return JsonResponse({"submissions": data})






# =====================================================================================================
@login_required
def admin_live_submissions(request, event_id):
    if not request.user.is_staff:
        return redirect("dashboard")

    event = get_object_or_404(Event, id=event_id)

    submissions = (
        UserChallenge.objects
        .filter(challenge__event=event)
        .select_related("user", "challenge")
        .order_by("-submitted_at")[:200]
    )

    return render(request, "administration/live_submissions.html", {
        "event": event,
        "submissions": submissions
    })






# ADD WAVE=============================================
from django.utils import timezone







@login_required
def admin_all_challenges(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    challenges = Challenge.objects.filter(
        event=event
    ).select_related("wave").order_by("wave__start_date", "points")

    return render(request, "administration/admin_all_challenges.html", {
        "event": event,
        "challenges": challenges
    })


# =====================================================================================================
# BULK USER IMPORT
# =====================================================================================================

@login_required
def admin_import_users(request):
    if not request.user.is_staff:
        messages.error(request, "❌ Access denied")
        return redirect("dashboard")

    if request.method == "POST":
        try:
            excel_file = request.FILES.get("file")
            event_id = request.POST.get("event_id")
            
            if not excel_file:
                messages.error(request, "❌ No file uploaded")
                return redirect("admin_import_users")

            # Load Workbook
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Prepare Output Excel
            out_wb = openpyxl.Workbook()
            out_ws = out_wb.active
            out_ws.append(["First Name", "Last Name", "Email", "Username", "Password", "Status"])

            created_count = 0
            event_obj = None

            if event_id:
                event_obj = Event.objects.get(id=event_id)

            # Iterate Rows (Skip Header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Assume columns: First Name, Last Name, Email (optional) -> indices 0, 1, 2
                if not row or not row[0]: # Skip empty rows
                    continue
                
                first_name = str(row[0]).strip()
                last_name = str(row[1]).strip()
                email = str(row[2]).strip() if len(row) > 2 and row[2] else ""

                # GENERATE CREDENTIALS
                safe_first = "".join(filter(str.isalnum, first_name)).lower()
                safe_last = "".join(filter(str.isalnum, last_name)).lower()
                
                base_username = f"{safe_first}.{safe_last}"
                username = base_username
                
                # Unique Username Logic
                counter = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}_{counter}"
                    counter += 1

                # Random Password
                alphabet = string.ascii_letters + string.digits
                password = ''.join(secrets.choice(alphabet) for i in range(10))

                # Create User
                try:
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                        email=email
                    )
                    created_count += 1
                    status = "Created"

                    # Grant Event Access
                    if event_obj:
                        access, _ = EventAccess.objects.get_or_create(
                            user=user, 
                            event=event_obj,
                            defaults={'is_registered': True}
                        )
                        if not access.is_registered:
                            access.is_registered = True
                            access.save()
                            
                        status += " + Event Access"

                except Exception as e:
                    status = f"Error: {str(e)}"
                    username = "N/A"
                    password = "N/A"

                out_ws.append([first_name, last_name, email, username, password, status])

            # Prepare Response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=generated_credentials.xlsx'
            out_wb.save(response)
            return response

        except Exception as e:
            messages.error(request, f"❌ Import failed: {str(e)}")
            return redirect("admin_import_users")

    # GET Request
    events = Event.objects.filter(status__in=['upcoming', 'live']).order_by('-created_at')
    return render(request, "administration/import_users.html", {
        "events": events
    })
