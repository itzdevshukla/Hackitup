from ctf.utils import encode_id, decode_id
from django.http import JsonResponse
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Min, Max
from datetime import datetime
from django.utils import timezone
from administration.models import Event, EventRole
from challenges.models import Challenge, UserChallenge, ChallengeHint, UserHint, ChallengeAttachment, ChallengeWave, WriteUp
from dashboard.models import EventAccess
import json
import openpyxl
import secrets
import string
from django.http import HttpResponse
from django.contrib.auth.hashers import make_password
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET

def is_admin(user, event_id=None):
    if not user.is_authenticated:
        return False
    # Superuser has access to everything
    if user.is_superuser:
        return True
    
    # If event_id is provided, check EventRole
    if event_id is not None:
        return EventRole.objects.filter(user=user, event_id=event_id).exists()
        
    # Generic admin endpoints (dashboard, user list, etc.) still require is_staff
    return user.is_staff

@login_required
def admin_dashboard_api(request):
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    total_users = User.objects.count()
    total_events = Event.objects.count()
    live_events = Event.objects.filter(status='live').count()
    upcoming_events = Event.objects.filter(status='upcoming').count()

    return JsonResponse({
        "stats": {
            "total_users": total_users,
            "total_events": total_events,
            "live_events": live_events,
            "upcoming_events": upcoming_events
        }
    })


@login_required
def admin_users_api(request):
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    users = User.objects.all().order_by('-date_joined')
    
    total_users = users.count()
    active_users = users.filter(is_active=True).count()
    admin_count = users.filter(is_staff=True).count()
    first_day = datetime.now().replace(day=1, hour=0, minute=0, second=0)
    new_users = users.filter(date_joined__gte=first_day).count()

    users_data = []
    for user in users:
        users_data.append({
            "id": encode_id(user.id),
            "username": user.username,
            "email": user.email,
            "is_active": user.is_active,
            "is_staff": user.is_staff,
            "date_joined": user.date_joined.strftime("%Y-%m-%d"),
        })

    return JsonResponse({
        "stats": {
            "total_users": total_users,
            "active_users": active_users,
            "admin_count": admin_count,
            "new_users": new_users,
        },
        "users": users_data
    })


@login_required
def admin_user_detail_api(request, user_id):
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        user = User.objects.get(id=user_id)
        
        # Get solved challenges
        solved_challenges = UserChallenge.objects.filter(user=user, is_correct=True).select_related('challenge', 'challenge__event')
        solved_data = []
        total_points = 0
        for sc in solved_challenges:
            challenge = sc.challenge
            event_name = challenge.event.event_name if challenge and challenge.event else "Unknown Event"
            
            solved_data.append({
                "challenge_id": encode_id(challenge.id if challenge else None),
                "title": challenge.title if challenge else "Unknown",
                "event": event_name,
                "points": challenge.points if challenge else 0,
                "submitted_at": sc.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if sc.submitted_at else None
            })
            if challenge and challenge.points:
                total_points += challenge.points
            
        # Get joined events
        joined_events = EventAccess.objects.filter(user=user).select_related('event')
        events_data = []
        for ae in joined_events:
            event = ae.event
            events_data.append({
                "id": encode_id(event.id if event else None),
                "name": event.event_name if event else "Unknown Event",
                "status": event.status if event and getattr(event, 'status', None) else "unknown"
            })

        return JsonResponse({
            "user": {
                "id": encode_id(user.id),
                "username": user.username,
                "email": user.email,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "is_active": user.is_active,
                "is_staff": user.is_staff,
                "date_joined": user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else None,
            },
            "stats": {
                "total_points": total_points,
                "challenges_solved": len(solved_data),
                "events_joined": len(events_data)
            },
            "solved_challenges": solved_data,
            "events": events_data
        })
    except User.DoesNotExist:
        return JsonResponse({"error": "User not found"}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def admin_events_api(request):
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    if request.user.is_staff or request.user.is_superuser:
        events = Event.objects.all().order_by('-created_at')
    else:
        events = Event.objects.filter(eventrole__user=request.user).order_by('-created_at')
    
    total_events = events.count()
    live_events = events.filter(status='live').count()
    upcoming_events = events.filter(status='upcoming').count()
    completed_events = events.filter(status='completed').count()

    events_data = []
    for event in events:
        events_data.append({
            "id": encode_id(event.id),
            "name": event.event_name,
            "venue": event.venue,
            "status": event.status,
            "access_code": event.access_code,
            "start_date": event.start_date.strftime("%Y-%m-%d") if event.start_date else None,
            "participants": EventAccess.objects.filter(event=event).count(),
            "is_approved": event.is_approved,
            "is_rejected": event.is_rejected
        })

    return JsonResponse({
        "stats": {
            "total_events": total_events,
            "live_events": live_events,
            "upcoming_events": upcoming_events,
            "completed_events": completed_events,
        },
        "events": events_data
    })


@login_required
@csrf_exempt
def admin_event_requests_api(request):
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
    
    requests = Event.objects.filter(is_approved=False, is_rejected=False).order_by('-created_at')
    
    requests_data = []
    for req in requests:
        requests_data.append({
            "id": encode_id(req.id),
            "event_name": req.event_name,
            "venue": req.venue,
            "description": req.description,
            "ctf_type": req.ctf_type,
            "max_participants": req.max_participants,
            "start_date": req.start_date.isoformat() if hasattr(req.start_date, 'isoformat') else req.start_date,
            "created_by": req.created_by.username,
        })
        
    return JsonResponse({"requests": requests_data})



@login_required
@csrf_exempt
def admin_approve_event_api(request, event_id):
    if request.method != 'POST':
        return JsonResponse({"error": "Method not allowed"}, status=405)
    
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
    
    try:
        event = Event.objects.get(id=event_id)
        event.is_approved = True
        event.is_rejected = False
        event.save()
        
        # Automatically make the requester an organizer
        EventRole.objects.get_or_create(
            event=event,
            user=event.created_by,
            defaults={'role': 'organizer'}
        )
        
        return JsonResponse({"success": True, "message": "Event approved successfully"})
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)


@login_required
@csrf_exempt
def admin_decline_event_api(request, event_id):
    if request.method != 'POST':
        return JsonResponse({"error": "Method not allowed"}, status=405)
    
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
    
    try:
        event = Event.objects.get(id=event_id)
        event.is_rejected = True
        event.is_approved = False
        event.save()
        return JsonResponse({"success": True, "message": "Event declined successfully"})
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)


@login_required
@csrf_exempt
def admin_add_event_api(request):
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Required fields
            required = ['event_name', 'venue', 'description', 'max_participants', 
                        'ctf_type', 'start_date', 'start_time', 'end_date', 'end_time']
            for field in required:
                if not data.get(field):
                    return JsonResponse({"error": f"Missing required field: {field}"}, status=400)
            
            # Create event
            event = Event.objects.create(
                event_name=data['event_name'],
                venue=data['venue'],
                description=data['description'],
                ctf_type=data['ctf_type'],
                max_participants=data['max_participants'],
                start_date=data['start_date'],
                start_time=data['start_time'],
                end_date=data['end_date'],
                end_time=data['end_time'],
                registration_start_date=data.get('reg_start_date') or None,
                registration_start_time=data.get('reg_start_time') or None,
                registration_end_date=data.get('reg_end_date') or None,
                registration_end_time=data.get('reg_end_time') or None,
                is_team_mode=str(data.get('is_team_mode')).lower() == 'true' if isinstance(data.get('is_team_mode'), str) else bool(data.get('is_team_mode')),
                max_team_size=int(data.get('max_team_size') or 4),
                created_by=request.user,
                is_approved=True 
            )
            
            # Auto-assign creator as 'organizer'
            EventRole.objects.create(event=event, user=request.user, role='organizer')
            
            return JsonResponse({
                "message": "Event created successfully",
                "event_id": encode_id(event.id),
                "access_code": event.access_code
            }, status=201)
            
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
            
    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
@csrf_exempt
def admin_edit_event_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)
        
    if request.method in ['PUT', 'POST']:
        try:
            data = json.loads(request.body)
            
            # Update fields if provided
            if 'event_name' in data: event.event_name = data['event_name']
            if 'venue' in data: event.venue = data['venue']
            if 'description' in data: event.description = data['description']
            if 'ctf_type' in data: event.ctf_type = data['ctf_type']
            if 'max_participants' in data: event.max_participants = data['max_participants']
            if 'access_code' in data: event.access_code = data['access_code']
            if 'start_date' in data: event.start_date = data['start_date']
            if 'start_time' in data: event.start_time = data['start_time']
            if 'end_date' in data: event.end_date = data['end_date']
            if 'end_time' in data: event.end_time = data['end_time']
            if 'reg_start_date' in data: event.registration_start_date = data.get('reg_start_date') or None
            if 'reg_start_time' in data: event.registration_start_time = data.get('reg_start_time') or None
            if 'reg_end_date' in data: event.registration_end_date = data.get('reg_end_date') or None
            if 'reg_end_time' in data: event.registration_end_time = data.get('reg_end_time') or None
            if 'is_team_mode' in data: 
                val = data.get('is_team_mode')
                event.is_team_mode = str(val).lower() == 'true' if isinstance(val, str) else bool(val)
            if 'max_team_size' in data: 
                event.max_team_size = int(data.get('max_team_size') or 4)
            
            event.save()
            return JsonResponse({"message": "Event updated successfully"})
            
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
            
    return JsonResponse({"error": "Method not allowed"}, status=405)

@login_required
@csrf_exempt
def admin_event_control_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)
        
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            now = timezone.localtime()
            
            if action == 'start_reg':
                event.registration_start_date = now.date()
                event.registration_start_time = now.time()
                
                # If there's no end date, or if the current end date/time is already in the past,
                # reset it to the overall event end date/time so the window actually stays open.
                current_reg_end = None
                if event.registration_end_date and event.registration_end_time:
                    current_reg_end = timezone.make_aware(
                        datetime.combine(event.registration_end_date, event.registration_end_time)
                    ) if timezone.is_naive(datetime.combine(event.registration_end_date, event.registration_end_time)) else datetime.combine(event.registration_end_date, event.registration_end_time)
                
                if not current_reg_end or current_reg_end <= now:
                    event.registration_end_date = event.end_date
                    event.registration_end_time = event.end_time
                    
                event.is_registration_paused = False
                event.save()
                return JsonResponse({"message": "Registration started manually"})
                
            elif action == 'stop_reg':
                past = timezone.now() - timezone.timedelta(seconds=1)
                event.registration_end_date = past.date()
                event.registration_end_time = past.time()
                event.is_registration_paused = True
                event.save()
                return JsonResponse({"message": "Registration stopped"})
                
            elif action == 'pause_event':
                event.is_paused = True
                event.save()
                return JsonResponse({"message": "Event paused"})
                
            elif action == 'resume_event':
                event.is_paused = False
                event.save()
                return JsonResponse({"message": "Event resumed"})
                
            elif action == 'end_event':
                past = timezone.localtime() - timezone.timedelta(seconds=1)
                event.end_date = past.date()
                event.end_time = past.time()
                event.is_paused = False
                event.save()
                return JsonResponse({"message": "Event ended permanently"})
                
            elif action == 'start_writeups':
                event.accepting_writeups = True
                event.save()
                return JsonResponse({"message": "Now accepting writeups"})
                
            elif action == 'stop_writeups':
                event.accepting_writeups = False
                event.save()
                return JsonResponse({"message": "Stopped accepting writeups"})
                
            else:
                return JsonResponse({"error": "Unknown action"}, status=400)
                
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
            
    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
def admin_event_detail_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
        
        # Get Challenges
        challenges = Challenge.objects.filter(event=event).order_by('points')
        challenges_data = []
        for c in challenges:
            challenges_data.append({
                "id": encode_id(c.id),
                "title": c.title,
                "description": c.description,
                "category": c.category,
                "points": c.points,
                "wave_id": encode_id(c.wave_id),
                "solves": UserChallenge.objects.filter(challenge=c, is_correct=True).count()
            })
            
        # Get Participants
        participants = EventAccess.objects.filter(event=event).select_related('user')
        participants_data = []
        for p in participants:
            participants_data.append({
                "id": encode_id(p.user.id),
                "username": p.user.username,
                "joined_at": p.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(p, 'created_at') else None
            })

        return JsonResponse({
            "event": {
                "id": encode_id(event.id),
                "name": event.event_name,
                "venue": event.venue,
                "status": event.status,
                "description": event.description,
                "ctf_type": event.ctf_type,
                "max_participants": event.max_participants,
                "rules": event.rules,
                "access_code": event.access_code,
                "is_approved": event.is_approved,
                "is_rejected": event.is_rejected,
                "is_paused": event.is_paused,
                "is_registration_paused": event.is_registration_paused,
                "is_registration_open": event.is_registration_open(),
                "accepting_writeups": event.accepting_writeups,
                "start_date": event.start_date.strftime("%Y-%m-%d") if event.start_date else None,
                "start_time": event.start_time.strftime("%H:%M") if event.start_time else None,
                "end_date": event.end_date.strftime("%Y-%m-%d") if event.end_date else None,
                "end_time": event.end_time.strftime("%H:%M") if event.end_time else None,
                "reg_start_date": event.registration_start_date.strftime("%Y-%m-%d") if event.registration_start_date else None,
                "reg_start_time": event.registration_start_time.strftime("%H:%M") if event.registration_start_time else None,
                "reg_end_date": event.registration_end_date.strftime("%Y-%m-%d") if event.registration_end_date else None,
                "reg_end_time": event.registration_end_time.strftime("%H:%M") if event.registration_end_time else None,
                "is_team_mode": event.is_team_mode,
                "max_team_size": event.max_team_size,
                "created_at": event.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(event, 'created_at') else None
            },
            "stats": {
                "total_challenges": len(challenges_data),
                "total_participants": len(participants_data)
            },
            "challenges": challenges_data,
            "participants": participants_data
        })
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)


@login_required
@csrf_exempt
def admin_user_writeups_api(request, event_id, user_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
        user = User.objects.get(id=user_id)
    except (Event.DoesNotExist, User.DoesNotExist):
        return JsonResponse({"error": "Event or User not found"}, status=404)
        
    if request.method == "GET":
        wus = WriteUp.objects.filter(user=user, challenge__event=event).select_related('challenge')
        data = []
        for w in wus:
            data.append({
                'challenge_id': encode_id(w.challenge.id),
                'challenge_title': w.challenge.title,
                'content': w.content,
                'submitted_at': w.challenge.created_at.isoformat() if hasattr(w.challenge, 'created_at') else None
            })
        return JsonResponse({'writeups': data})
        
    return JsonResponse({"error": "Method not allowed"}, status=405)

@csrf_exempt
@login_required
def admin_update_event_rules_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)
        
    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
            event.rules = data.get('rules', '')
            event.save()
            return JsonResponse({"message": "Rules updated successfully"})
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
            
    return JsonResponse({"error": "Method not allowed"}, status=405)

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@login_required
def admin_delete_user_api(request, user_id):
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    if request.method == 'DELETE':
        try:
            user = User.objects.get(id=user_id)
            if user.id == request.user.id:
                return JsonResponse({"error": "Cannot delete your own account"}, status=400)
            user.delete()
            return JsonResponse({"message": "User deleted successfully"})
        except User.DoesNotExist:
            return JsonResponse({"error": "User not found"}, status=404)
            
    return JsonResponse({"error": "Method not allowed"}, status=405)

@csrf_exempt
@login_required
def admin_delete_event_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    if request.method == 'DELETE':
        try:
            event = Event.objects.get(id=event_id)
            event.delete()
            return JsonResponse({"message": "Event deleted successfully"})
        except Event.DoesNotExist:
            return JsonResponse({"error": "Event not found"}, status=404)
            
    return JsonResponse({"error": "Method not allowed"}, status=405)

@csrf_exempt
@login_required
def admin_create_challenge_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
    
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)

    if request.method == "POST":
        try:
            data = request.POST
            title = data.get("title", "").strip()
            description = data.get("description", "").strip()
            category = data.get("category", "").strip()
            difficulty = data.get("difficulty", "easy").strip()
            points_raw = data.get("points", 0)
            flag = data.get("flag", "").strip()
            flag_format = data.get("flag_format", "Hack!tUp{...}").strip()
            url = data.get("url", "").strip()
            wave_id = data.get("wave_id", "").strip()

            wave = None
            if wave_id:
                try:
                    wave = ChallengeWave.objects.get(id=int(wave_id), event=event)
                except (ChallengeWave.DoesNotExist, ValueError):
                    pass

            if not title or not description or not category or not flag:
                return JsonResponse({"error": "title, description, category, and flag are required."}, status=400)

            try:
                points = int(points_raw)
            except (ValueError, TypeError):
                points = 0

            challenge = Challenge.objects.create(
                event=event,
                wave=wave,
                title=title,
                description=description,
                category=category,
                difficulty=difficulty,
                points=points,
                flag=flag,   # will be hashed by model.save()
                flag_format=flag_format or "Hack!tUp{...}",
                url=url or None,
                author=request.user
            )

            # Handle Hints
            hints_raw = data.get("hints", "[]")
            try:
                hints_list = json.loads(hints_raw)
                for h in hints_list:
                    content = str(h.get("content", "")).strip()
                    try:
                        cost = int(h.get("cost", 0))
                    except (ValueError, TypeError):
                        cost = 0
                    if content:
                        ChallengeHint.objects.create(challenge=challenge, content=content, cost=cost)
            except (json.JSONDecodeError, TypeError):
                pass

            # Handle file attachments
            for key, uploaded_file in request.FILES.items():
                if key.startswith("new_files"):
                    ChallengeAttachment.objects.create(challenge=challenge, file=uploaded_file)

            return JsonResponse({"message": "Challenge created successfully", "id": encode_id(challenge.id)}, status=201)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required
def admin_challenge_detail_api(request, event_id, challenge_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        challenge = Challenge.objects.get(id=challenge_id, event_id=event_id)
    except Challenge.DoesNotExist:
        return JsonResponse({"error": "Challenge not found"}, status=404)
        
    if request.method == "GET":
        return JsonResponse({
            "id": encode_id(challenge.id),
            "title": challenge.title,
            "description": challenge.description,
            "category": challenge.category,
            "points": challenge.points,
            "has_flag": bool(challenge.flag),
            "flag_format": challenge.flag_format,
            "url": challenge.url,
            "wave_id": encode_id(challenge.wave_id),
            "wave_name": challenge.wave.name if challenge.wave else None,
            "hints": [{"content": h.content, "cost": h.cost} for h in challenge.hints.all()],
            "attachments": [{"id": encode_id(a.id), "file_name": a.file.name.split('/')[-1], "file_url": a.file.url} for a in challenge.attachments.all()],
            "solves": UserChallenge.objects.filter(challenge=challenge, is_correct=True).count()
        })
        
    elif request.method == "POST" or request.method == "PUT":
        # We'll use POST or PUT with multipart/form-data to support file uploads.
        # However, Django's PUT via fetch doesn't populate request.POST automatically for multipart. 
        # A common workaround is to send a POST request with an overridden _method='PUT' or just accept POST for edits.
        # Let's handle both. If it's pure JSON, we use request.body. If it's multipart, we use request.POST.
        
        try:
            # Check if it's a file deletion action first
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                if data.get("action") == "delete_file":
                    file_id = data.get("file_id")
                    if file_id:
                        ChallengeAttachment.objects.filter(id=file_id, challenge=challenge).delete()
                        return JsonResponse({"message": "File deleted successfully"})
            
            # For standard updates (which might include files), we expect multipart/form-data.
            # In Django, if a fetch uses POST with FormData, data is in request.POST and files in request.FILES.
            data = request.POST if request.POST else json.loads(request.body)
            
            challenge.title = data.get("title", challenge.title)
            challenge.description = data.get("description", challenge.description)
            challenge.category = data.get("category", challenge.category)
            challenge.points = data.get("points", challenge.points)
            challenge.flag_format = data.get("flag_format", challenge.flag_format)
            challenge.url = data.get("url", challenge.url)

            # Handle wave assignment
            wave_id_raw = data.get("wave_id", "")
            if wave_id_raw == "" or wave_id_raw is None:
                challenge.wave = None
            else:
                try:
                    challenge.wave = ChallengeWave.objects.get(id=int(wave_id_raw), event_id=event_id)
                except (ChallengeWave.DoesNotExist, ValueError):
                    pass  # keep existing wave if invalid
            
            # Handle Hints Parsing (We might receive a JSON stringified array of dicts)
            if "hints" in data:
                hints_data = data.get("hints")
                # If it arrived via multipart POST, it might be a JSON string
                if isinstance(hints_data, str):
                    try:
                        hints_data = json.loads(hints_data)
                    except json.JSONDecodeError:
                        hints_data = []

                if isinstance(hints_data, list):
                    challenge.hints.all().delete()
                    for h in hints_data:
                        content = str(h.get("content", "")).strip()
                        cost_val = h.get("cost", 0)
                        try:
                            cost = int(cost_val)
                        except (ValueError, TypeError):
                            cost = 0
                        if content:
                            ChallengeHint.objects.create(challenge=challenge, content=content, cost=cost)

            # Handle File Uploads
            for key, uploaded_file in request.FILES.items():
                if key.startswith('new_files'):
                    ChallengeAttachment.objects.create(challenge=challenge, file=uploaded_file)
            
            # Handle Flag updating
            new_flag = data.get("flag", "").strip()
            if new_flag:
                challenge.flag = make_password(new_flag)
                
            challenge.save()
            return JsonResponse({"message": "Challenge updated successfully"})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": str(e)}, status=400)
            
    elif request.method == "DELETE":
        challenge.delete()
        return JsonResponse({"message": "Challenge deleted successfully"})
        
    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required
def admin_waves_api(request, event_id):
    """List all waves for an event (GET) or create a new wave (POST)."""
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)

    if request.method == "GET":
        waves = event.waves.all().order_by('order', 'created_at')
        return JsonResponse({"waves": [
            {
                "id": encode_id(w.id),
                "name": w.name,
                "order": w.order,
                "is_active": w.is_active,
                "challenge_count": w.challenges.count(),
            } for w in waves
        ]})

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            name = data.get("name", "").strip()
            order = int(data.get("order", 0))
            if not name:
                return JsonResponse({"error": "Wave name is required"}, status=400)
            wave = ChallengeWave.objects.create(event=event, name=name, order=order)
            return JsonResponse({"id": encode_id(wave.id), "name": wave.name, "order": wave.order, "is_active": wave.is_active, "challenge_count": 0}, status=201)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required
def admin_wave_detail_api(request, event_id, wave_id):
    """Toggle active (PUT), or delete a wave (DELETE)."""
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
    try:
        wave = ChallengeWave.objects.get(id=wave_id, event_id=event_id)
    except ChallengeWave.DoesNotExist:
        return JsonResponse({"error": "Wave not found"}, status=404)

    if request.method == "PUT":
        try:
            data = json.loads(request.body)
            was_active = wave.is_active
            
            if "is_active" in data:
                wave.is_active = bool(data["is_active"])
            if "name" in data:
                name = data["name"].strip()
                if name:
                    wave.name = name
            if "order" in data:
                wave.order = int(data["order"])
            wave.save()
            
            # Auto-generate announcement when a wave is unpaused
            if not was_active and wave.is_active:
                from challenges.models import Announcement
                Announcement.objects.create(
                    event_id=event_id,
                    title=f"Wave Released: {wave.name}",
                    content=f"Attention hackers! The challenges for {wave.name} are now LIVE. Good luck!",
                    type="info",
                    created_by=request.user
                )
                
            return JsonResponse({"id": encode_id(wave.id), "name": wave.name, "is_active": wave.is_active, "order": wave.order})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    if request.method == "DELETE":
        wave.delete()
        return JsonResponse({"message": "Wave deleted successfully"})

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required
def admin_wave_challenges_api(request, event_id, wave_id):
    """GET: all event challenges with assignment status. PUT: bulk-assign challenge IDs to this wave."""
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
    try:
        wave = ChallengeWave.objects.get(id=wave_id, event_id=event_id)
    except ChallengeWave.DoesNotExist:
        return JsonResponse({"error": "Wave not found"}, status=404)

    if request.method == "GET":
        all_challenges = Challenge.objects.filter(event_id=event_id).order_by('category', 'title')
        return JsonResponse({"challenges": [
            {
                "id": encode_id(c.id),
                "title": c.title,
                "category": c.category,
                "points": c.points,
                "wave_id": encode_id(c.wave_id),
                "wave_name": c.wave.name if c.wave else None,
            } for c in all_challenges
        ]})

    if request.method == "PUT":
        try:
            data = json.loads(request.body)
            raw_ids = data.get("challenge_ids", [])
            # Decode hashids to real integer PKs
            challenge_ids = [decode_id(hid) for hid in raw_ids if decode_id(hid) is not None]
            # Unassign any current challenges from this wave that aren't in the new list
            Challenge.objects.filter(wave=wave).exclude(id__in=challenge_ids).update(wave=None)
            # Assign selected challenges to this wave
            Challenge.objects.filter(id__in=challenge_ids, event_id=event_id).update(wave=wave)
            return JsonResponse({"message": "Challenges assigned successfully", "count": len(challenge_ids)})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)



@csrf_exempt
@login_required
def admin_import_users_api(request):
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)


    if request.method == "POST":
        try:
            excel_file = request.FILES.get("file")
            event_id = request.POST.get("event_id")
            
            if not excel_file:
                return JsonResponse({"error": "No file uploaded"}, status=400)

            # Load Workbook
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Prepare Output Excel
            out_wb = openpyxl.Workbook()
            out_ws = out_wb.active
            out_ws.append(["First Name", "Last Name", "Email", "Username", "Password", "Status"])

            event_obj = None
            if event_id and event_id != '' and event_id != 'None':
                decoded_event_id = decode_id(event_id)
                if decoded_event_id is None:
                    return JsonResponse({"error": "Invalid event ID"}, status=400)
                event_obj = Event.objects.get(id=decoded_event_id)

            created_count = 0
            # Iterate Rows (Skip Header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: # Skip empty rows
                    continue
                
                first_name = str(row[0]).strip()
                last_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                email = str(row[2]).strip() if len(row) > 2 and row[2] else ""

                # GENERATE CREDENTIALS
                safe_first = "".join(filter(str.isalnum, first_name)).lower()
                safe_last = "".join(filter(str.isalnum, last_name)).lower()
                
                base_username = f"{safe_first}.{safe_last}"
                if not base_username or base_username == ".":
                    base_username = "user"
                username = base_username
                
                # Unique Username Logic
                counter = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}_{counter}"
                    counter += 1

                # Random Password
                alphabet = string.ascii_letters + string.digits
                password = ''.join(secrets.choice(alphabet) for i in range(10))

                # Create User
                try:
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                        email=email
                    )
                    created_count += 1
                    status = "Created"

                    # Grant Event Access
                    if event_obj:
                        access, _ = EventAccess.objects.get_or_create(
                            user=user, 
                            event=event_obj,
                            defaults={'is_registered': True}
                        )
                        if not access.is_registered:
                            access.is_registered = True
                            access.save()
                            
                        status += " + Event Access"

                except Exception as e:
                    status = f"Error: {str(e)}"
                    username = "N/A"
                    password = "N/A"

                out_ws.append([first_name, last_name, email, username, password, status])

            # Prepare Response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="generated_credentials.xlsx"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            out_wb.save(response)
            return response

        except Exception as e:
            return JsonResponse({"error": f"Import failed: {str(e)}"}, status=500)

    # GET Request: Generate active events for modal dropdown
    events = Event.objects.filter(status__in=['upcoming', 'live']).order_by('-created_at')
    events_data = [{"id": encode_id(e.id), "name": e.event_name} for e in events]
    return JsonResponse({"events": events_data})

@csrf_exempt
@login_required
def admin_event_teams_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
        if not event.is_team_mode:
            return JsonResponse({"error": "Event is not in Team Mode"}, status=400)
            
        from teams.models import Team
        from dashboard.models import EventAccess
        teams = Team.objects.filter(event=event).prefetch_related('members__user').order_by('-created_at')
        
        # We need to know which users are banned in this event
        access_records = EventAccess.objects.filter(event=event)
        banned_user_ids = set(access_records.filter(is_banned=True).values_list('user_id', flat=True))
        
        teams_data = []
        for team in teams:
            members = []
            for tm in team.members.all():
                members.append({
                    "user_id": encode_id(tm.user.id),
                    "username": tm.user.username,
                    "is_captain": tm.user.id == team.captain_id,
                    "is_banned": tm.user.id in banned_user_ids
                })
            
            teams_data.append({
                "team_id": encode_id(team.id),
                "team_name": team.name,
                "invite_code": team.invite_code,
                "captain": team.captain.username,
                "member_count": len(members),
                "created_at": team.created_at.strftime("%Y-%m-%d %H:%M:%S") if team.created_at else None,
                "members": members,
                "is_banned": all(m['is_banned'] for m in members) if members else False
            })
            
        return JsonResponse({
            "event_name": event.event_name,
            "teams": teams_data,
            "total_teams": len(teams_data)
        })
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)


@csrf_exempt
@login_required
def admin_event_participants_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
        registrations = EventAccess.objects.filter(event=event).select_related('user').order_by('-granted_at')
        
        user_to_team = {}
        if event.is_team_mode:
            from teams.models import Team
            teams = Team.objects.filter(event=event).prefetch_related('members__user')
            for team in teams:
                for member in team.members.all():
                    user_to_team[member.user.id] = team.name

        participants_data = []
        for reg in registrations:
            participants_data.append({
                "id": encode_id(reg.user.id),
                "username": reg.user.username,
                "email": reg.user.email,
                "joined_at": reg.granted_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(reg, 'granted_at') and reg.granted_at else None,
                "is_registered": reg.is_registered,
                "is_banned": reg.is_banned,
                "team_name": user_to_team.get(reg.user.id, "No Team") if event.is_team_mode else None
            })
            
        return JsonResponse({
            "event_name": event.event_name,
            "is_team_mode": event.is_team_mode,
            "participants": participants_data,
            "total_participants": len(participants_data)
        })
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)

@csrf_exempt
@login_required
def admin_event_leaderboard_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)

    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)

    # ── Always read is_team_mode from the DB — never trust client input ──────
    if event.is_team_mode:
        from teams.models import Team, TeamChallenge, TeamMember

        teams = Team.objects.filter(event=event).prefetch_related(
            'members__user', 'solves__challenge'
        )

        team_list = []
        for team in teams:
            solves = team.solves.all()
            total_points = sum(s.challenge.points for s in solves)
            last_solve_obj = solves.order_by('-solved_at').first()
            last_solve_str = last_solve_obj.solved_at.strftime("%Y-%m-%d %H:%M:%S") if last_solve_obj else None

            members = []
            for m in team.members.all():
                # Check if this member is banned
                try:
                    from dashboard.models import EventAccess
                    access = EventAccess.objects.get(event=event, user=m.user)
                    is_banned = access.is_banned
                except:
                    is_banned = False

                members.append({
                    "user_id": encode_id(m.user.id),
                    "username": m.user.username,
                    "is_captain": m.user.id == team.captain_id,
                    "is_banned": is_banned,
                })

            team_list.append({
                "team_id": encode_id(team.id),
                "team_name": team.name,
                "captain": team.captain.username,
                "member_count": len(members),
                "members": members,
                "total_points": total_points,
                "solves": solves.count(),
                "last_solve": last_solve_str,
                "_last_solve_raw": last_solve_obj.solved_at if last_solve_obj else None,
            })

        # Sort: points desc, last_solve asc (teams with no solves go last)
        from django.utils import timezone as tz
        team_list.sort(key=lambda t: (
            -t["total_points"],
            t["_last_solve_raw"] if t["_last_solve_raw"] else tz.now()
        ))

        for rank, team in enumerate(team_list, start=1):
            team["rank"] = rank
            del team["_last_solve_raw"]  # internal field, don't expose

        return JsonResponse({
            "event_name": event.event_name,
            "is_team_mode": True,
            "leaderboard": team_list,
        })

    # ── Individual mode (unchanged) ──────────────────────────────────────────
    leaderboard = list(
        UserChallenge.objects
        .filter(
            challenge__event=event,
            is_correct=True,
            user__eventaccess__event=event,
            user__eventaccess__is_banned=False
        )
        .values("user__id", "user__username")
        .annotate(
            total_points=Sum("challenge__points"),
            solves=Count("id"),
            first_solve=Min("submitted_at"),
            last_solve=Max("submitted_at")
        )
    )

    user_hints = (
        UserHint.objects
        .filter(hint__challenge__event=event)
        .values("user__id")
        .annotate(total_deduction=Sum("hint__cost"))
    )
    hint_deductions = {item["user__id"]: item["total_deduction"] for item in user_hints}

    for entry in leaderboard:
        deduction = hint_deductions.get(entry["user__id"], 0)
        entry["total_points"] = max(0, entry["total_points"] - deduction)

    leaderboard.sort(key=lambda x: (-x["total_points"], x["last_solve"]))

    leaderboard_data = []
    for rank, entry in enumerate(leaderboard, start=1):
        leaderboard_data.append({
            "rank": rank,
            "user_id": encode_id(entry["user__id"]),
            "username": entry["user__username"],
            "total_points": entry["total_points"],
            "solves": entry["solves"],
            "last_solve": entry["last_solve"].strftime("%Y-%m-%d %H:%M:%S") if entry["last_solve"] else None
        })

    return JsonResponse({
        "event_name": event.event_name,
        "is_team_mode": False,
        "leaderboard": leaderboard_data,
    })

@csrf_exempt
@login_required
def admin_event_submissions_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
        
        # Check flag visibility permissions
        can_see_flag = request.user.is_superuser or (event.created_by == request.user)
        if not can_see_flag:
            can_see_flag = EventRole.objects.filter(event=event, user=request.user, role='organizer').exists()
        
        # User to Team Mapping if team mode
        user_to_team = {}
        if event.is_team_mode:
            from teams.models import Team
            teams = Team.objects.filter(event=event).prefetch_related('members__user')
            for t in teams:
                for member in t.members.all():
                    user_to_team[member.user.id] = {
                        "name": t.name,
                        "id": encode_id(t.id)
                    }

        submissions = (
            UserChallenge.objects
            .filter(challenge__event=event)
            .select_related("user", "challenge")
            .order_by("-submitted_at")[:200]
        )
        
        submissions_data = []
        for s in submissions:
            flag_val = s.submitted_flag if (can_see_flag or not s.is_correct) else "CORRECT"
            
            sub_dict = {
                "id": encode_id(s.id),
                "user_id": encode_id(s.user.id),
                "username": s.user.username,
                "challenge_id": encode_id(s.challenge.id),
                "challenge_title": s.challenge.title,
                "flag": flag_val,
                "is_correct": s.is_correct,
                "submitted_at": s.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if s.submitted_at else None
            }
            
            if event.is_team_mode:
                team_info = user_to_team.get(s.user.id, {"name": "No Team", "id": None})
                sub_dict["team_name"] = team_info["name"]
                sub_dict["team_id"] = team_info["id"]

            submissions_data.append(sub_dict)
            
        return JsonResponse({
            "event_name": event.event_name,
            "is_team_mode": event.is_team_mode,
            "submissions": submissions_data
        })
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)

@csrf_exempt
@login_required
def admin_user_event_submissions_api(request, event_id, user_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
        user = User.objects.get(id=user_id)
        
        # Check flag visibility permissions
        can_see_flag = request.user.is_superuser or (event.created_by == request.user)
        if not can_see_flag:
            can_see_flag = EventRole.objects.filter(event=event, user=request.user, role='organizer').exists()

        try:
            from dashboard.models import EventAccess
            access = EventAccess.objects.get(event=event, user=user)
            is_banned = access.is_banned
        except:
            is_banned = False

        submissions = (
            UserChallenge.objects
            .filter(challenge__event=event, user=user)
            .select_related("challenge")
            .order_by("-submitted_at")
        )
        
        submissions_data = []
        for s in submissions:
            flag_val = s.submitted_flag if (can_see_flag or not s.is_correct) else "CORRECT"
            submissions_data.append({
                "id": encode_id(s.id),
                "challenge_title": s.challenge.title,
                "flag": flag_val,
                "is_correct": s.is_correct,
                "submitted_at": s.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if s.submitted_at else None
            })
            
        hints = (
            UserHint.objects
            .filter(hint__challenge__event=event, user=user)
            .select_related("hint__challenge")
            .order_by("-unlocked_at")
        )
        hints_data = []
        for h in hints:
            hints_data.append({
                "id": encode_id(h.id),
                "challenge_title": h.hint.challenge.title,
                "cost": h.hint.cost,
                "unlocked_at": h.unlocked_at.strftime("%Y-%m-%d %I:%M %p") if h.unlocked_at else None
            })
            
        return JsonResponse({
            "event_name": event.event_name,
            "username": user.username,
            "is_banned": is_banned,
            "submissions": submissions_data,
            "hints_taken": hints_data
        })
    except (Event.DoesNotExist, User.DoesNotExist):
        return JsonResponse({"error": "Event or User not found"}, status=404)


@csrf_exempt
@login_required
def admin_team_submissions_api(request, event_id, team_id):
    """
    Returns all UserChallenge submissions made by any member of a given team.
    Server-side guard: event.is_team_mode is read from DB — no client flag trusted.
    """
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)

    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)

    # Server-side guard — only valid for team-mode events
    if not event.is_team_mode:
        return JsonResponse({"error": "This event is not in team mode"}, status=400)

    from teams.models import Team, TeamMember
    try:
        team = Team.objects.get(id=team_id, event=event)
    except Team.DoesNotExist:
        return JsonResponse({"error": "Team not found"}, status=404)

    # Flag visibility — same rules as individual submissions
    can_see_flag = request.user.is_superuser or (event.created_by == request.user)
    if not can_see_flag:
        can_see_flag = EventRole.objects.filter(event=event, user=request.user, role='organizer').exists()

    # Collect all member user IDs for this team
    member_user_ids = TeamMember.objects.filter(team=team).values_list('user_id', flat=True)

    submissions = (
        UserChallenge.objects
        .filter(challenge__event=event, user_id__in=member_user_ids)
        .select_related("user", "challenge")
        .order_by("-submitted_at")
    )

    submissions_data = []
    for s in submissions:
        flag_val = s.submitted_flag if (can_see_flag or not s.is_correct) else "CORRECT"
        submissions_data.append({
            "id": encode_id(s.id),
            "username": s.user.username,
            "user_id": encode_id(s.user.id),
            "challenge_title": s.challenge.title,
            "flag": flag_val,
            "is_correct": s.is_correct,
            "submitted_at": s.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if s.submitted_at else None,
        })

    return JsonResponse({
        "event_name": event.event_name,
        "team_name": team.name,
        "submissions": submissions_data,
    })


@csrf_exempt
@login_required
def admin_toggle_ban_team_api(request, event_id, team_id):
    """
    Bans or unbans all members of a specific team.
    If any member is unbanned, it will ban all members. 
    If all members are banned, it will unban all members.
    """
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Unauthorized"}, status=401)

    if request.method == 'POST':
        try:
            event = Event.objects.get(id=event_id)

            # Allow: superuser, staff, organizer, or event_admin for this event
            is_super = request.user.is_superuser or request.user.is_staff
            has_role = EventRole.objects.filter(
                event=event,
                user=request.user,
                role__in=['organizer', 'admin']
            ).exists()

            if not is_super and not has_role:
                return JsonResponse({"error": "Forbidden — insufficient permissions to ban teams"}, status=403)

            from teams.models import Team
            team = Team.objects.get(id=team_id, event=event)

            from dashboard.models import EventAccess
            member_users = team.members.values_list('user', flat=True)
            access_records = EventAccess.objects.filter(event=event, user__in=member_users)

            # Determine target state: if ANY member is not banned, we BAN all of them.
            # Only if ALL are already banned, we UNBAN them.
            target_ban_state = any(not access.is_banned for access in access_records)

            # Update all member accesses
            access_records.update(is_banned=target_ban_state)

            return JsonResponse({
                "message": f"Team {'banned' if target_ban_state else 'unbanned'} successfully",
                "is_banned": target_ban_state
            })
        except (Event.DoesNotExist, Team.DoesNotExist):
            return JsonResponse({"error": "Event or Team record not found"}, status=404)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required
def admin_delete_team_api(request, event_id, team_id):
    """
    Deletes a specified team and removes the members from the EventAccess 
    so they can join other teams or the event again.
    """
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Unauthorized"}, status=401)

    if request.method == 'DELETE':
        try:
            event = Event.objects.get(id=event_id)

            # Allow: superuser, staff, organizer, or event_admin
            is_super = request.user.is_superuser or request.user.is_staff
            has_role = EventRole.objects.filter(
                event=event,
                user=request.user,
                role__in=['organizer', 'admin']
            ).exists()

            if not is_super and not has_role:
                return JsonResponse({"error": "Forbidden — insufficient permissions"}, status=403)

            from teams.models import Team
            team = Team.objects.get(id=team_id, event=event)
            
            # Remove event access for all team members so they can join/create new teams
            from dashboard.models import EventAccess
            member_users = team.members.values_list('user', flat=True)
            EventAccess.objects.filter(event=event, user__in=member_users).delete()

            team_name = team.name
            team.delete()

            return JsonResponse({"message": f"Team '{team_name}' and its members' event access deleted successfully"})

        except (Event.DoesNotExist, Team.DoesNotExist):
            return JsonResponse({"error": "Event or Team not found"}, status=404)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required
def admin_toggle_ban_participant_api(request, event_id, user_id):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Unauthorized"}, status=401)

    if request.method == 'POST':
        try:
            event = Event.objects.get(id=event_id)

            # Allow: superuser, staff, organizer, or event_admin for this event
            is_super = request.user.is_superuser or request.user.is_staff
            has_role = EventRole.objects.filter(
                event=event,
                user=request.user,
                role__in=['organizer', 'admin']
            ).exists()

            if not is_super and not has_role:
                return JsonResponse({"error": "Forbidden — insufficient permissions to ban users"}, status=403)

            user = User.objects.get(id=user_id)
            access = EventAccess.objects.get(event=event, user=user)

            # Toggle the ban status
            access.is_banned = not access.is_banned
            access.save()

            # The user's score will dynamically drop from the leaderboard because the leaderboard
            # query now strictly filters out users where `is_banned=True`.

            return JsonResponse({
                "message": f"User {'banned' if access.is_banned else 'unbanned'} successfully",
                "is_banned": access.is_banned
            })
        except (Event.DoesNotExist, User.DoesNotExist, EventAccess.DoesNotExist):
            return JsonResponse({"error": "Event, User, or Access record not found"}, status=404)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required
def admin_export_event_data_api(request, event_id):
    if not is_admin(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
        
        # Prepare Workbook and Sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Event Data Export"

        # Headers Setup - Big Event Title on top
        ws.merge_cells('A1:E1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"USER DETAILS : {event.event_name.upper()}"
        from openpyxl.styles import Font, Alignment
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Leave row 2 blank or details
        ws.merge_cells('A2:E2')
        detail_cell = ws.cell(row=2, column=1)
        detail_cell.value = f"Total Participants: {EventAccess.objects.filter(event=event).count()}"
        detail_cell.font = Font(italic=True)
        detail_cell.alignment = Alignment(horizontal='center')
        
        # Blank row
        ws.append([])

        # Data Headers
        headers = ["Rank", "Username", "Email", "Total Points", "Challenges Solved"]
        ws.append(headers)
        
        # Style Data Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.font = Font(bold=True)

        # Get all registered participants for the event
        all_participants = EventAccess.objects.filter(event=event).select_related('user')
        
        # Get Leaderboard logic for those who actually scored
        leaderboard_data = (
            UserChallenge.objects
            .filter(
                challenge__event=event,
                is_correct=True,
                user__eventaccess__event=event
            )
            .values("user__id")
            .annotate(
                total_points=Sum("challenge__points"),
                solves=Count("id"),
                last_solve=Max("submitted_at")
            )
        )
        
        # Create a dictionary for quick lookup of scores
        scores_dict = {
            item["user__id"]: {
                "total_points": item["total_points"],
                "solves": item["solves"],
                "last_solve": item["last_solve"]
            } for item in leaderboard_data
        }
        
        # Combine participants with their scores
        combined_data = []
        for p in all_participants:
            user_id = p.user.id
            score_info = scores_dict.get(user_id, {"total_points": 0, "solves": 0, "last_solve": None})
            
            combined_data.append({
                "username": p.user.username,
                "email": p.user.email,
                "total_points": score_info["total_points"] or 0,
                "solves": score_info["solves"] or 0,
                "last_solve": score_info["last_solve"]
            })
            
        # Sort the combined data: first by points (descending), then by last_solve time (ascending)
        # Note: Users with no solves have last_solve as None, handle that in sorting
        combined_data.sort(
            key=lambda x: (
                -x["total_points"], 
                x["last_solve"] if x["last_solve"] else timezone.now() # Push non-solvers down
            )
        )
        
        # Append ranked data to worksheet
        for rank, entry in enumerate(combined_data, start=1):
            ws.append([
                rank,
                entry["username"],
                entry["email"],
                entry["total_points"],
                entry["solves"]
            ])

        # Prepare Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{event.event_name.replace(" ", "_").lower()}_export.xlsx"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        wb.save(response)
        return response

    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)

@csrf_exempt
@login_required
def admin_event_roles_api(request, event_id):
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)
        
    is_organizer_or_super = request.user.is_superuser or EventRole.objects.filter(event=event, user=request.user, role='organizer').exists()
    
    if request.method == "GET":
        roles = EventRole.objects.filter(event=event).select_related('user')
        roles_data = [
            {
                "id": encode_id(r.id),
                "user_id": encode_id(r.user.id),
                "username": r.user.username,
                "email": r.user.email,
                "role": r.role,
                "created_at": r.created_at.strftime("%Y-%m-%d %H:%M:%S")
            } for r in roles
        ]
        return JsonResponse({
            "roles": roles_data,
            "current_user_role": "organizer" if is_organizer_or_super else "admin",
            "current_user_id": encode_id(request.user.id)
        })
        
    if not is_organizer_or_super:
        return JsonResponse({"error": "Only organizers can modify roles"}, status=403)
        
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            username = data.get("username", "").strip()
            role = data.get("role", "admin").strip()
            
            if role not in ['organizer', 'admin']:
                return JsonResponse({"error": "Invalid role type"}, status=400)
                
            user_to_add = User.objects.get(username=username)
            event_role, created = EventRole.objects.get_or_create(event=event, user=user_to_add, defaults={'role': role})
            
            if not created:
                event_role.role = role
                event_role.save()
                
            return JsonResponse({"message": f"Successfully assigned {role} role to {username}"}, status=201)
        except User.DoesNotExist:
            return JsonResponse({"error": "User does not exist. They must register an account first."}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
            
    if request.method == "DELETE":
        try:
            data = json.loads(request.body)
            role_id_hash = data.get("role_id")
            role_id = decode_id(role_id_hash)
            if role_id is None:
                return JsonResponse({"error": "Invalid role ID"}, status=400)
            
            role_to_delete = EventRole.objects.get(id=role_id, event=event)
            if role_to_delete.user == request.user and role_to_delete.role == 'organizer':
                return JsonResponse({"error": "Cannot remove yourself as an organizer."}, status=400)
                
            role_to_delete.delete()
            return JsonResponse({"message": "Role removed successfully"})
        except EventRole.DoesNotExist:
            return JsonResponse({"error": "Role mapping not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
            
    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required
def admin_test_challenges_list_api(request, event_id):
    """
    Returns all challenges for the specific event to Admins for testing,
    regardless of waves or active status.
    """
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)
        
    if request.method == "GET":
        challenges = Challenge.objects.filter(event=event).order_by('category', 'title')
        
        challenges_data = []
        for c in challenges:
            challenges_data.append({
                "id": encode_id(c.id),
                "title": c.title,
                "description": c.description,
                "category": c.category,
                "points": c.points,
                "wave_id": encode_id(c.wave_id),
                "wave_name": c.wave.name if c.wave else None,
                "flag_format": c.flag_format,
                "solves_count": 0, # Placeholder for UI
                "is_solved": False, # Admins are testing, so always false initially
                "first_blood": None,
                "author": c.author.username if c.author else "Unknown",
                "difficulty": c.difficulty,
                "url": c.url,
                "files": [{"url": f.file.url} for f in c.attachments.all()],
                "hints": [{"id": encode_id(h.id), "cost": h.cost, "is_unlocked": True, "content": h.content} for h in c.hints.all()] # Admins always see hints
            })
            
        return JsonResponse({
            "event": event.event_name,
            "challenges": challenges_data
        })
        
    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@login_required
def admin_test_challenge_flag_api(request, challenge_id):
    """
    Validates a flag entry against a challenge for testing purposes only.
    It does not record UserChallenge models or update scores.
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Method not allowed"}, status=405)
        
    try:
        challenge = Challenge.objects.get(id=challenge_id)
    except Challenge.DoesNotExist:
        return JsonResponse({"error": "Challenge not found"}, status=404)
        
    # Security: Verify the user testing it is an admin for this specific event
    if not is_admin(request.user, event_id=challenge.event.id):
        return JsonResponse({"error": "Forbidden"}, status=403)
        
    try:
        data = json.loads(request.body)
        submitted_flag = data.get('flag', '').strip()
        
        from django.contrib.auth.hashers import check_password
        
        # Verify Flag
        is_correct = check_password(submitted_flag, challenge.flag) or (submitted_flag == challenge.flag)
        
        if is_correct:
            return JsonResponse({"success": True, "message": "Valid Flag - System Compromised (Test)"})
        else:
            return JsonResponse({"success": False, "error": "Invalid Flag (Test)"}, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@login_required
@require_POST
def admin_create_announcement_api(request, event_id):
    """
    Creates a new Announcement for the given event.
    Only accessible by an Admin of the event.
    """
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden: Admin access only"}, status=403)
        
    try:
        event = Event.objects.get(id=event_id)
        data = json.loads(request.body)
        title = data.get('title', '').strip()
        content = data.get('content', '').strip()
        announcement_type = data.get('type', 'info').strip().lower()
        
        if not title or not content:
            return JsonResponse({"error": "Title and content are required."}, status=400)
            
        valid_types = [t[0] for t in __import__('challenges.models', fromlist=['Announcement']).Announcement.TYPE_CHOICES]
        if announcement_type not in valid_types:
            announcement_type = 'info'
            
        from challenges.models import Announcement
        
        announcement = Announcement.objects.create(
            event=event,
            title=title,
            content=content,
            type=announcement_type,
            created_by=request.user
        )
        
        return JsonResponse({
            "success": True, 
            "message": "Announcement created.",
            "announcement": {
                "id": encode_id(announcement.id),
                "title": announcement.title,
                "content": announcement.content,
                "type": announcement.type,
                "created_at": announcement.created_at.isoformat()
            }
        })
        
    except Event.DoesNotExist:
        return JsonResponse({"error": "Event not found"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON payload"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@login_required
@require_POST
def admin_delete_announcement_api(request, event_id, announcement_id):
    """
    Deletes an Announcement for the given event.
    Only accessible by an Admin of the event.
    """
    if not is_admin(request.user, event_id=event_id):
        return JsonResponse({"error": "Forbidden: Admin access only"}, status=403)
        
    try:
        from challenges.models import Announcement
        announcement = Announcement.objects.get(id=announcement_id, event_id=event_id)
        announcement.delete()
        
        return JsonResponse({"success": True, "message": "Announcement deleted successfully."})
        
    except Announcement.DoesNotExist:
        return JsonResponse({"error": "Announcement not found."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
